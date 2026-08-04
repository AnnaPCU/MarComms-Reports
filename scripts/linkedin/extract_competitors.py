# Extrae el benchmark de COMPETIDORES por cuenta, para un mes.
# Soporta ambos formatos:
#   - unified:  <base>/*.xlsx con hoja que contiene 'compet'
#   - crudo:    <base>/<carpeta cuenta>/*competitor*.xlsx (hoja COMPETITORS)
# Salida: JSON {acc: [{name, fol, nfol, eng, posts, own}]}
import os, sys, glob, json
import warnings
warnings.filterwarnings('ignore')
import openpyxl
from extract_unified import MATCHERS
from extract_raw import FOLDER_MATCHERS

# Nombre "propio" esperado por cuenta (para marcar la fila own en el set).
OWN = {
    'cul': ['control union latinoam', 'controlunion latinoam', 'control union services'],
    'cue': ['control union espa'],
    'cup': ['control union portugal'],
    'cun': ['control union norte'],
    'cuna': ['control union north'],
    'ps':  ['peterson solutions'],
    'pia': ['peterson solutions (iberia', 'peterson solutions iberia'],
    'tlr': ['tlr'],
    'bel': ['biomass energy lab'],
}

def rows_of(ws):
    return [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]

def parse_comp(rows, acc):
    # Header: Página / Total de seguidores / Nuevos seguidores / Total de interacciones... / Total de publicaciones
    hi = None
    for i, r in enumerate(rows[:4]):
        vals = [str(c).strip() if c is not None else '' for c in r]
        if 'Página' in vals:
            hi = i; hdr = vals; break
    if hi is None:
        return []
    ci = {n: hdr.index(n) for n in hdr if n}
    out = []
    for r in rows[hi + 1:]:
        name = r[ci['Página']] if ci.get('Página') is not None else None
        if not name:
            continue
        def g(colname):
            j = ci.get(colname)
            v = r[j] if j is not None and j < len(r) else 0
            return int(v) if isinstance(v, (int, float)) else 0
        nm = str(name).strip()
        own = any(p in nm.lower() for p in OWN[acc])
        out.append(dict(name=nm, fol=g('Total de seguidores'), nfol=g('Nuevos seguidores'),
                        eng=g('Total de interacciones en la publicación'), posts=g('Total de publicaciones'), own=own))
    return out

def from_unified(base):
    xlsx = [f for f in os.listdir(base) if f.lower().endswith('.xlsx')]
    res = {}
    for acc, pred in MATCHERS.items():
        hits = [f for f in xlsx if pred(f)]
        if len(hits) != 1:
            res[acc] = []; continue
        wb = openpyxl.load_workbook(os.path.join(base, hits[0]), data_only=True)
        comp = [s for s in wb.sheetnames if 'compet' in s.lower()]
        res[acc] = parse_comp(rows_of(wb[comp[0]]), acc) if comp else []
    return res

def from_raw(base):
    res = {}
    dirs = [d for d in os.listdir(base) if os.path.isdir(os.path.join(base, d))]
    for acc, pred in FOLDER_MATCHERS.items():
        hits = [d for d in dirs if pred(d.lower())]
        if len(hits) != 1:
            res[acc] = []; continue
        files = glob.glob(os.path.join(base, hits[0], '*competitor*.xlsx'))
        if not files:
            res[acc] = []; continue
        wb = openpyxl.load_workbook(files[0], data_only=True)
        res[acc] = parse_comp(rows_of(wb[wb.sheetnames[0]]), acc)
    return res

if __name__ == '__main__':
    base, out = sys.argv[1], sys.argv[2]
    has_dirs = any(os.path.isdir(os.path.join(base, d)) for d in os.listdir(base))
    res = from_raw(base) if has_dirs else from_unified(base)
    for acc, comp in res.items():
        ownrow = next((c for c in comp if c['own']), None)
        print(f"  {acc:5} {len(comp):2d} páginas · own={'SÍ' if ownrow else 'NO'}", file=sys.stderr)
    json.dump(res, open(out, 'w'), ensure_ascii=False, indent=1)
    print("wrote", out)
