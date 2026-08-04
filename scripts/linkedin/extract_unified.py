import openpyxl, os, json, sys, re

# Cada cuenta se resuelve por un predicado sobre el nombre de archivo
# (los nombres cambian levemente mes a mes: mayúsculas, "global", etc.).
def _match(name, has, hasnt=()):
    n=name.lower()
    return all(h in n for h in has) and not any(x in n for x in hasnt)
MATCHERS={
 'cul': lambda n:_match(n,['latinoameric']),
 'cue': lambda n:_match(n,['espana']),
 'cup': lambda n:_match(n,['portugal']),
 'cun': lambda n:_match(n,['norte']),
 'cuna':lambda n:_match(n,['north']),
 'ps':  lambda n:_match(n,['peterson'],['iberia','america']),
 'pia': lambda n:_match(n,['peterson','iberia']),
 'tlr': lambda n:_match(n,['tlr']) or _match(n,['laborator']),
 'bel': lambda n:_match(n,['biomass']),
}
def resolve_files(base):
    import os as _os
    xlsx=[f for f in _os.listdir(base) if f.lower().endswith('.xlsx')]
    out={}
    for acc,pred in MATCHERS.items():
        hits=[f for f in xlsx if pred(f)]
        if len(hits)!=1:
            raise SystemExit(f"Resolución ambigua para {acc}: {hits} en {base}")
        out[acc]=hits[0]
    return out

def _index_map(wb):
    # La hoja "Índice" mapea 'Hoja original' -> 'Nombre de la hoja en el consolidado'.
    idx=[s for s in wb.sheetnames if 'ndice' in s.lower()]
    if not idx: return {}
    rows=list(wb[idx[0]].iter_rows(values_only=True))
    hdr=[str(c).strip() if c is not None else '' for c in rows[0]]
    try:
        ci_o=hdr.index('Hoja original'); ci_c=hdr.index('Nombre de la hoja en el consolidado')
    except ValueError:
        return {}
    m={}
    for r in rows[1:]:
        if r[ci_o] and r[ci_c]: m[str(r[ci_o]).strip().lower()]=str(r[ci_c]).strip()
    return m

def sheet(wb, sub):
    # 1) Vía Índice por nombre de hoja ORIGINAL (estable entre meses).
    ORIG={'indicador':'indicadores','todas':'todas las publicaciones',
          'nuevos':'nuevos seguidores','datos de vi':'datos de visitantes'}
    m=_index_map(wb)
    key=ORIG.get(sub)
    if key and key in m:
        target=m[key].strip()
        actual=next((s for s in wb.sheetnames if s.strip()==target), None)
        if actual:
            return list(wb[actual].iter_rows(values_only=True))
    # 2) Fallback: substring sobre el nombre consolidado.
    for s in wb.sheetnames:
        if sub in s.lower(): return list(wb[s].iter_rows(values_only=True))
    return []
def find_header(rows, must):
    for i,r in enumerate(rows[:6]):
        vals=[str(c).strip() if c is not None else '' for c in r]
        if all(any(m==v for v in vals) for m in must): return i,vals
    return None,None
def col(h,n):
    if not h: return None
    for i,v in enumerate(h):
        if v==n: return i
    return None
def ssum(rows,hi,ci):
    if ci is None or hi is None: return 0
    t=0
    for r in rows[hi+1:]:
        v=r[ci] if ci<len(r) else None
        if isinstance(v,(int,float)): t+=v
    return t

def clean_title(t):
    if not t: return ''
    lines=[re.sub(r'\s+',' ', l.replace('\xa0',' ')).strip() for l in str(t).split('\n')]
    lines=[l for l in lines if l]
    out=''
    for l in lines:
        out = l if not out else out+' '+l
        if len(out)>=40: break  # juntar líneas cortas (fechas, etc.) hasta tener contexto
    return out[:110]

def extract(base):
    res={}
    FILES=resolve_files(base)
    for acc,fn in FILES.items():
        wb=openpyxl.load_workbook(os.path.join(base,fn), data_only=True)
        ind=sheet(wb,'indicador')
        hi,hdr=find_header(ind,['Fecha','Impresiones (totales)','Clics (totales)'])
        imp=ssum(ind,hi,col(hdr,'Impresiones (totales)')); clk=ssum(ind,hi,col(hdr,'Clics (totales)'))
        rea=ssum(ind,hi,col(hdr,'Reacciones (total)')); com=ssum(ind,hi,col(hdr,'Comentarios (totales)')); shr=ssum(ind,hi,col(hdr,'Veces compartido (total)'))
        eng=clk+rea+com+shr; er=round(eng/imp*100,2) if imp else 0
        fr=sheet(wb,'nuevos'); fhi,fhdr=find_header(fr,['Fecha','Total de seguidores']); fol=ssum(fr,fhi,col(fhdr,'Total de seguidores'))
        vr=sheet(wb,'datos de vi'); vhi,vhdr=find_header(vr,['Fecha','Visitantes únicos en total (total)']); vis=ssum(vr,vhi,col(vhdr,'Visitantes únicos en total (total)'))
        # posts
        pr=sheet(wb,'todas')
        posts=[]
        if pr:
            phi,phdr=find_header(pr,['Título de la publicación','Impresiones','Clics'])
            if phi is not None:
                ci_t=col(phdr,'Título de la publicación'); ci_u=col(phdr,'Enlace de la publicación')
                ci_imp=col(phdr,'Impresiones'); ci_clk=col(phdr,'Clics'); ci_er=col(phdr,'Tasa de interacción')
                ci_tp=col(phdr,'Tipo de contenido'); ci_pub=col(phdr,'Tipo de publicación')
                for r in pr[phi+1:]:
                    if not r or not r[ci_t]: continue
                    imp_p=r[ci_imp] if ci_imp is not None and ci_imp<len(r) else 0
                    if not isinstance(imp_p,(int,float)): imp_p=0
                    clk_p=r[ci_clk] if ci_clk is not None and ci_clk<len(r) else 0
                    if not isinstance(clk_p,(int,float)): clk_p=0
                    er_p=r[ci_er] if ci_er is not None and ci_er<len(r) else 0
                    er_p=round(er_p*100,2) if isinstance(er_p,(int,float)) else 0
                    tp=r[ci_tp] if ci_tp is not None and ci_tp<len(r) else ''
                    pub=r[ci_pub] if ci_pub is not None and ci_pub<len(r) else ''
                    tp=(tp or '').strip() or ('Orgánico' if (pub or '').strip().lower().startswith('org') else (pub or 'Orgánico'))
                    posts.append(dict(t=clean_title(r[ci_t]),
                                      url=(r[ci_u] if ci_u is not None and ci_u<len(r) else '') or '',
                                      imp=int(imp_p), clk=int(clk_p), er=er_p, tp=tp))
        posts.sort(key=lambda p:p['imp'], reverse=True)
        res[acc]=dict(imp=imp,clk=clk,er=er,vis=vis,fol=fol,posts=posts[:6])
    return res

if __name__=='__main__':
    base=sys.argv[1]; out=sys.argv[2]
    res=extract(base)
    json.dump(res, open(out,'w'), ensure_ascii=False, indent=1)
    print("wrote", out)
